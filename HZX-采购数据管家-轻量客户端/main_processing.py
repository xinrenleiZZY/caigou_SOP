"""
HZX-采购数据管家 - 核心处理引擎
从原main.py提取，适配GUI调用
"""
import sys, os, re, shutil, time, zipfile, tempfile
from copy import copy
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pypinyin

# 修正配置导入
if getattr(sys, 'frozen', False):
    _BASE = os.path.dirname(sys.executable)
else:
    _BASE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, _BASE)

def log_callback(msg):
    """日志回调，GUI可覆盖此函数"""
    print(f"[{time.strftime('%H:%M:%S')}] {msg}")

# ========== 工具函数 ==========
def fix_autofilter(file_path):
    tdir = tempfile.mkdtemp()
    with zipfile.ZipFile(file_path, 'r') as z:
        z.extractall(tdir)
    fixed = 0
    for root, dirs, files in os.walk(tdir):
        for f in files:
            fp = os.path.join(root, f)
            try:
                content = open(fp, 'rb').read().decode('utf-8')
            except:
                continue
            if '<autoFilter' in content:
                old = content
                content = re.sub(r'<autoFilter[^>]*>.*?</autoFilter>', '', content, flags=re.DOTALL)
                content = re.sub(r'<autoFilter[^>]*/>', '', content)
                if content != old:
                    open(fp, 'wb').write(content.encode('utf-8'))
                    fixed += 1
    os.remove(file_path)
    with zipfile.ZipFile(file_path, 'w', zipfile.ZIP_DEFLATED) as z:
        for root, dirs, files in os.walk(tdir):
            for f in files:
                z.write(os.path.join(root, f), os.path.relpath(os.path.join(root, f), tdir))
    shutil.rmtree(tdir)
    return fixed

def copy_style(src, dst):
    if src.has_style:
        dst.font = copy(src.font)
        dst.border = copy(src.border)
        dst.fill = copy(src.fill)
        dst.number_format = copy(src.number_format)
        dst.alignment = copy(src.alignment)

def detect_latest_common_month(source_file, year1, year2, log=log_callback):
    wb = load_workbook(source_file, read_only=True, data_only=True)
    sheet_names = set(wb.sheetnames)
    wb.close()
    p1 = re.compile(rf'^{re.escape(str(year1))}\.(\d+)$')
    p2 = re.compile(rf'^{re.escape(str(year2))}\.(\d+)$')
    m1s, m2s = set(), set()
    for s in sheet_names:
        m = p1.match(s)
        if m: m1s.add(int(m.group(1)))
        m = p2.match(s)
        if m: m2s.add(int(m.group(1)))
    common = m1s & m2s
    if not common:
        m1 = max(m1s) if m1s else None
        m2 = max(m2s) if m2s else None
        log(f"警告: 未找到共有月份, {year1}最新={m1}, {year2}最新={m2}")
        return m1, m2
    max_month = max(common)
    log(f"共有月份: {sorted(common)}, 取最新: {max_month}")
    return max_month, max_month

def calculate_all_data(source_file, sheet1, sheet2):
    wb = load_workbook(source_file, read_only=True, data_only=True)
    result = {}
    for sn in [sheet1, sheet2]:
        ws = wb[sn]; pd = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            p = str(row[3]).strip() if row[3] else ''
            if p not in ['Amazon', 'temu', 'jit']: continue
            qty = row[7] or 0; amt = row[10] or 0
            ch = str(row[11]).strip() if row[11] else ''
            cs = str(row[15]).strip() if row[15] else ''
            if p not in pd:
                pd[p] = {'total_qty': 0, 'total_amt': 0, 'gz_qty': 0, 'gz_amt': 0,
                         'yw_qty': 0, 'yw_amt': 0, 'customs_qty': 0, 'customs_amt': 0}
            d = pd[p]; d['total_qty'] += qty; d['total_amt'] += amt
            if ch == '广州': d['gz_qty'] += qty; d['gz_amt'] += amt
            elif ch == '义乌': d['yw_qty'] += qty; d['yw_amt'] += amt
            if cs == '是': d['customs_qty'] += qty; d['customs_amt'] += amt
        result[sn] = pd
    wb.close()
    return result

def calculate_channel_data(source_file, sheet1, sheet2):
    wb = load_workbook(source_file, read_only=True, data_only=True)
    result = {}
    for sn in [sheet1, sheet2]:
        ws = wb[sn]; gzq, gza, ywq, ywa = 0, 0, 0, 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            p = str(row[3]).strip() if row[3] else ''
            if p != 'Amazon': continue
            qty = row[7] or 0; amt = row[10] or 0
            ch = str(row[11]).strip() if row[11] else ''
            if ch == '广州': gzq += qty; gza += amt
            elif ch == '义乌': ywq += qty; ywa += amt
        result[sn] = {'gz_qty': gzq, 'gz_amt': round(gza, 2),
                       'yw_qty': ywq, 'yw_amt': round(ywa, 2),
                       'total_qty': gzq + ywq, 'total_amt': round(gza + ywa, 2)}
    wb.close()
    return result

def get_next_channel_columns(ws):
    col = 2
    while True:
        val = ws.cell(row=1, column=col).value
        if val is not None and str(val).replace('.', '').isdigit():
            col += 2
        else:
            break
    return col, col + 1

def get_next_platform_column(ws):
    col = 2
    while True:
        val = ws.cell(row=3, column=col).value
        if val is not None:
            col += 1
        else:
            break
    return col

def adjust_all_formulas(ws, insert_at, count=2):
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if v and isinstance(v, str) and v.startswith('='):
                def repl(m):
                    pre = m.group(1); rn = m.group(2)
                    dollar = '$' if rn.startswith('$') else ''
                    clean = rn.lstrip('$')
                    num = int(clean)
                    if num >= insert_at:
                        return f'{pre}{dollar}{num + count}'
                    return m.group(0)
                cell.value = re.sub(r'(\$?[A-Z]+\$?)(\d+)', repl, v)

def _pinyin_key(name):
    """拼音排序键：(空白) 排最后，中文按拼音排序"""
    if name == '(空白)':
        return (1, '')
    try:
        py = ''.join(pypinyin.lazy_pinyin(name))
        return (0, py)
    except:
        return (0, name)

def build_amazon_hierarchy(ws, max_row):
    rows = []
    for r in range(2, max_row + 1):
        platform = ws.cell(row=r, column=4).value
        if platform and str(platform).strip() == 'Amazon':
            qty = ws.cell(row=r, column=8).value or 0
            amt = ws.cell(row=r, column=11).value or 0
            channel = str(ws.cell(row=r, column=12).value or '').strip()
            rm = str(ws.cell(row=r, column=13).value or '').strip()
            rn = str(ws.cell(row=r, column=14).value or '').strip()
            ro = str(ws.cell(row=r, column=15).value or '').strip()
            cs = str(ws.cell(row=r, column=16).value or '').strip()
            rows.append({'qty': qty, 'amt': amt, 'channel': channel,
                         'reason_m': rm, 'reason_n': rn, 'reason_o': ro, 'customs': cs})
    total_qty = sum(r['qty'] for r in rows)
    total_amt = sum(r['amt'] for r in rows)
    result = [{'level': 'platform', 'channel': '', 'reason_m': '', 'reason_n': '',
               'reason_o': '', 'customs': '', 'qty': total_qty, 'amt': round(total_amt, 2)}]
    channels = {}
    for row in rows:
        channels.setdefault(row['channel'], []).append(row)
    for ch_name in ['广州', '义乌']:
        if ch_name not in channels: continue
        ch_rows = channels[ch_name]
        ch_qty = sum(r['qty'] for r in ch_rows)
        ch_amt = sum(r['amt'] for r in ch_rows)
        result.append({'level': 'channel', 'channel': ch_name, 'reason_m': '',
                       'reason_n': '', 'reason_o': '', 'customs': '',
                       'qty': ch_qty, 'amt': ch_amt})
        rms = {}
        for row in ch_rows:
            rm_key = row['reason_m'] if row['reason_m'] else '(空白)'
            rms.setdefault(rm_key, []).append(row)
        for rm_name, rm_rows in sorted(rms.items(), key=lambda x: _pinyin_key(x[0])):
            rm_qty = sum(r['qty'] for r in rm_rows)
            rm_amt = sum(r['amt'] for r in rm_rows)
            result.append({'level': 'reason_m', 'channel': '', 'reason_m': rm_name,
                           'reason_n': '', 'reason_o': '', 'customs': '',
                           'qty': rm_qty, 'amt': rm_amt})
            rns = {}
            for row in rm_rows:
                rn_key = row['reason_n'] if row['reason_n'] else '(空白)'
                rns.setdefault(rn_key, []).append(row)
            for rn_name, rn_rows in sorted(rns.items(), reverse=True):
                rn_qty = sum(r['qty'] for r in rn_rows)
                rn_amt = sum(r['amt'] for r in rn_rows)
                result.append({'level': 'reason_n', 'channel': '', 'reason_m': '',
                               'reason_n': rn_name, 'reason_o': '', 'customs': '',
                               'qty': rn_qty, 'amt': rn_amt})
                ros = {}
                for row in rn_rows:
                    ro_key = row['reason_o'] if row['reason_o'] else '(空白)'
                    ros.setdefault(ro_key, []).append(row)
                for ro_name, ro_rows in sorted(ros.items(), reverse=True):
                    ro_qty = sum(r['qty'] for r in ro_rows)
                    ro_amt = sum(r['amt'] for r in ro_rows)
                    result.append({'level': 'reason_o', 'channel': '', 'reason_m': '',
                                   'reason_n': '', 'reason_o': ro_name, 'customs': '',
                                   'qty': ro_qty, 'amt': ro_amt})
                    cs_map = {}
                    for row in ro_rows:
                        cs_key = row['customs'] if row['customs'] else '(空白)'
                        cs_map.setdefault(cs_key, []).append(row)
                    for cs_name, cs_rows in sorted(cs_map.items(), reverse=True):
                        cs_qty = sum(r['qty'] for r in cs_rows)
                        cs_amt = sum(r['amt'] for r in cs_rows)
                        result.append({'level': 'customs', 'channel': '', 'reason_m': '',
                                       'reason_n': '', 'reason_o': '', 'customs': cs_name,
                                       'qty': cs_qty, 'amt': cs_amt})
    result.append({'level': 'total', 'channel': '总计', 'reason_m': '',
                   'reason_n': '', 'reason_o': '', 'customs': '',
                   'qty': total_qty, 'amt': round(total_amt, 2)})
    return result

# ========== 主处理函数 ==========
def run_processing(source_file, output_file, year1, year2, enable_border=False, log=log_callback):
    """执行全部5个任务的主处理函数"""
    start_time = time.time()
    
    try:
        log("=" * 50)
        log(f"HZX-采购数据管家 V1.0.0")
        log(f"源文件: {source_file}")
        log(f"输出文件: {output_file}")
        log(f"年份: {year1}/{year2}")
        log("=" * 50)
        
        # 检测月份
        month1, month2 = detect_latest_common_month(source_file, year1, year2, log)
        if month1 is None or month2 is None:
            log("错误: 未找到数据工作表")
            return False
        
        sheet1 = f"{year1}.{month1}"
        sheet2 = f"{year2}.{month2}"
        
        # 输出文件名加入月份
        if "{month}" in output_file:
            output_file = output_file.replace("{month}", f"{month1}")
        
        log(f"检测到月份: {sheet1}, {sheet2}")
        
        # 复制并清理
        SHEET_CH_AMAZON = f"{year2}年出货渠道对比-Amazon"
        SHEET_CH_COMPARE = f"{year1}年、{year2}年出货渠道对比"
        SHEET_3PLATFORM = f"{year1}年、{year2}年采购量和采购总额对比-三平台"
        SHEET_CUSTOMS = f"{year1}年、{year2}年报关占比表--Amazon"
        SHEETS_KEEP = [sheet1, sheet2, SHEET_CH_AMAZON, SHEET_CH_COMPARE, SHEET_3PLATFORM, SHEET_CUSTOMS]
        
        shutil.copy2(source_file, output_file)
        log("已复制源文件")
        fixed = fix_autofilter(output_file)
        log(f"已修复 {fixed} 个autofilter")
        
        wb = load_workbook(output_file)
        sheets_to_remove = [s for s in wb.sheetnames if s not in SHEETS_KEEP]
        for s in sheets_to_remove:
            del wb[s]
        wb.save(output_file)
        wb.close()
        log(f"已清理工作表，保留6个")
        
        # 计算数据
        all_data = calculate_all_data(source_file, sheet1, sheet2)
        channel_data = calculate_channel_data(source_file, sheet1, sheet2)
        ch1 = channel_data[sheet1]; ch2 = channel_data[sheet2]
        log(f"{sheet1} Amazon: 广州={ch1['gz_qty']}, 义乌={ch1['yw_qty']}, 合计={ch1['total_qty']}")
        log(f"{sheet2} Amazon: 广州={ch2['gz_qty']}, 义乌={ch2['yw_qty']}, 合计={ch2['total_qty']}")
        
        amz_pd1 = all_data[sheet1].get('Amazon', {})
        amz_pd2 = all_data[sheet2].get('Amazon', {})
        customs1 = {'total_qty': amz_pd1.get('total_qty', 0), 'total_amt': amz_pd1.get('total_amt', 0),
                    'customs_qty': amz_pd1.get('customs_qty', 0), 'customs_amt': amz_pd1.get('customs_amt', 0)}
        customs2 = {'total_qty': amz_pd2.get('total_qty', 0), 'total_amt': amz_pd2.get('total_amt', 0),
                    'customs_qty': amz_pd2.get('customs_qty', 0), 'customs_amt': amz_pd2.get('customs_amt', 0)}
        
        # 重新打开
        wb = load_workbook(output_file)
        
        # ===== 任务1: 出货渠道对比 =====
        ws = wb[SHEET_CH_AMAZON]
        last_row = ws.max_row; gz_row = last_row + 1; yw_row = last_row + 2
        for c in range(1, 9):
            copy_style(ws.cell(row=8, column=c), ws.cell(row=gz_row, column=c))
            copy_style(ws.cell(row=9, column=c), ws.cell(row=yw_row, column=c))
        ws.cell(row=gz_row, column=1).value = month2
        ws.cell(row=gz_row, column=2).value = '广州'
        ws.cell(row=gz_row, column=3).value = ch2['gz_qty']
        ws.cell(row=gz_row, column=4).value = f'=C{gz_row}/$E${gz_row}'
        ws.cell(row=gz_row, column=5).value = f'=C{gz_row}+C{yw_row}'
        ws.cell(row=gz_row, column=6).value = ch2['gz_amt']
        ws.cell(row=gz_row, column=7).value = f'=F{gz_row}/$H${gz_row}'
        ws.cell(row=gz_row, column=8).value = f'=F{gz_row}+F{yw_row}'
        ws.cell(row=yw_row, column=2).value = '义乌'
        ws.cell(row=yw_row, column=3).value = ch2['yw_qty']
        ws.cell(row=yw_row, column=4).value = f'=C{yw_row}/$E${gz_row}'
        ws.cell(row=yw_row, column=6).value = ch2['yw_amt']
        ws.cell(row=yw_row, column=7).value = f'=F{yw_row}/$H${gz_row}'
        ws.merge_cells(start_row=gz_row, start_column=1, end_row=yw_row, end_column=1)
        ws.merge_cells(start_row=gz_row, start_column=5, end_row=yw_row, end_column=5)
        ws.merge_cells(start_row=gz_row, start_column=8, end_row=yw_row, end_column=8)
        from openpyxl.styles import Alignment
        al = Alignment(horizontal='center', vertical='center')
        ws.cell(row=gz_row, column=1).alignment = al
        ws.cell(row=gz_row, column=5).alignment = al
        ws.cell(row=gz_row, column=8).alignment = al
        log("任务1: 出货渠道对比完成")
        
        # ===== 任务2: 出货渠道对比列 =====
        ws = wb[SHEET_CH_COMPARE]
        val_col, pct_col = get_next_channel_columns(ws)
        from openpyxl.utils import get_column_letter
        clv = get_column_letter(val_col); clp = get_column_letter(pct_col)
        ws.column_dimensions[clv].width = ws.column_dimensions['B'].width or 9
        ws.column_dimensions[clp].width = ws.column_dimensions['C'].width or 9
        for sr, dr in [(1,1),(2,2),(3,3),(4,4),(7,7),(8,8),(9,9),(10,10),(13,13),(14,14),(15,15)]:
            copy_style(ws.cell(row=sr, column=2), ws.cell(row=dr, column=val_col))
            copy_style(ws.cell(row=sr, column=3), ws.cell(row=dr, column=pct_col))
        ws.merge_cells(start_row=1, start_column=val_col, end_row=1, end_column=pct_col)
        ws.cell(row=1, column=val_col).value = f'{year1}.{month1}'; ws.cell(row=1, column=val_col).alignment = al
        ws.cell(row=2, column=val_col).value = ch1['gz_qty']; ws.cell(row=2, column=pct_col).value = f'={clv}2/{clv}4'
        ws.cell(row=3, column=val_col).value = ch1['yw_qty']; ws.cell(row=3, column=pct_col).value = f'={clv}3/{clv}4'
        ws.cell(row=4, column=val_col).value = f'={clv}2+{clv}3'; ws.cell(row=4, column=pct_col).value = f'={clv}4/{clv}4'
        ws.merge_cells(start_row=7, start_column=val_col, end_row=7, end_column=pct_col)
        ws.cell(row=7, column=val_col).value = f'{year2}.{month2}'; ws.cell(row=7, column=val_col).alignment = al
        ws.cell(row=8, column=val_col).value = ch2['gz_qty']; ws.cell(row=8, column=pct_col).value = f'={clv}8/{clv}$10'
        ws.cell(row=9, column=val_col).value = ch2['yw_qty']; ws.cell(row=9, column=pct_col).value = f'={clv}9/{clv}$10'
        ws.cell(row=10, column=val_col).value = f'={clv}8+{clv}9'; ws.cell(row=10, column=pct_col).value = f'={clv}10/{clv}$10'
        ws.merge_cells(start_row=13, start_column=val_col, end_row=13, end_column=pct_col)
        ws.cell(row=13, column=val_col).value = month2; ws.cell(row=13, column=val_col).alignment = al
        ws.cell(row=14, column=val_col).value = f'={clv}8-{clv}2'; ws.cell(row=14, column=pct_col).value = f'={clp}8-{clp}2'
        ws.cell(row=15, column=val_col).value = f'={clv}9-{clv}3'; ws.cell(row=15, column=pct_col).value = f'={clp}9-{clp}3'
        log(f"任务2: 出货渠道对比列 {clv}-{clp} 完成")
        
        # ===== 任务3: 三平台 =====
        ws = wb[SHEET_3PLATFORM]
        new_col = get_next_platform_column(ws)
        cl = get_column_letter(new_col)
        ws.column_dimensions[cl].width = ws.column_dimensions['B'].width or 12
        platform_positions = {'Amazon': 1, 'temu': 18, 'jit': 35}
        for platform, start in platform_positions.items():
            pd1 = all_data.get(sheet1, {}).get(platform, {})
            pd2 = all_data.get(sheet2, {}).get(platform, {})
            q1 = pd1.get('total_qty', 0); a1 = round(pd1.get('total_amt', 0), 2)
            q2 = pd2.get('total_qty', 0); a2 = round(pd2.get('total_amt', 0), 2)
            qd = q2 - q1; ad = round(a2 - a1, 2)
            for sr, dr in [(start+2, start+2),(start+3, start+3),(start+4, start+4),
                            (start+7, start+7),(start+8, start+8),(start+9, start+9),
                            (start+12, start+12),(start+13, start+13),(start+14, start+14)]:
                copy_style(ws.cell(row=sr, column=2), ws.cell(row=dr, column=new_col))
            ws.cell(row=start+2, column=new_col).value = f'{year1}.{month1}'
            ws.cell(row=start+3, column=new_col).value = q1; ws.cell(row=start+4, column=new_col).value = a1
            ws.cell(row=start+7, column=new_col).value = f'{year2}.{month2}'
            ws.cell(row=start+8, column=new_col).value = q2; ws.cell(row=start+9, column=new_col).value = a2
            ws.cell(row=start+12, column=new_col).value = month2
            ws.cell(row=start+13, column=new_col).value = qd; ws.cell(row=start+14, column=new_col).value = ad
        log(f"任务3: 三平台 {cl} 列完成")
        
        # ===== 任务4: 报关占比表 =====
        ws = wb[SHEET_CUSTOMS]
        def find_month_end(ws, mv):
            for r in range(1, ws.max_row + 1):
                if ws.cell(row=r, column=1).value == mv:
                    return r + 2
            return None
        m1_label = int(f'{2000+year1}{month1:02d}')
        m2_label = int(f'{2000+year2}{month2:02d}')
        pos1 = find_month_end(ws, int(f'{2000+year1}{month1-1:02d}'))
        pos2 = find_month_end(ws, int(f'{2000+year2}{month2-1:02d}'))
        if pos1 is None: pos1 = 14
        if pos2 is None: pos2 = pos1 + 16
        
        for idx, (ins, label, cd) in enumerate([(pos1, m1_label, customs1), (pos2 + 2, m2_label, customs2)]):
            ws.insert_rows(ins, 2)
            adjust_all_formulas(ws, ins, 2)
            tr = [13, 14]
            for t_idx in [0, 1]:
                for c in range(1, 7):
                    copy_style(ws.cell(row=tr[t_idx], column=c), ws.cell(row=ins + t_idx, column=c))
            ws.cell(row=ins, column=1).value = label
            ws.cell(row=ins, column=3).value = cd['total_qty']
            ws.cell(row=ins, column=4).value = round(cd['total_amt'], 2)
            ws.cell(row=ins+1, column=1).value = label
            ws.cell(row=ins+1, column=2).value = '是'
            ws.cell(row=ins+1, column=3).value = cd['customs_qty']
            ws.cell(row=ins+1, column=4).value = round(cd['customs_amt'], 2)
            ws.cell(row=ins+1, column=5).value = f'=C{ins+1}/C{ins}'
            ws.cell(row=ins+1, column=6).value = f'=D{ins+1}/D{ins}'
        log("任务4: 报关占比表完成")
        
        # ===== 任务5: 数据分析表 =====
        for sn, ssn in [(f'{year1}.{month1}数据分析', sheet1), (f'{year2}.{month2}数据分析', sheet2)]:
            ws_src = wb[ssn]
            hierarchy = build_amazon_hierarchy(ws_src, ws_src.max_row)
            if sn in wb.sheetnames: del wb[sn]
            ws = wb.create_sheet(title=sn)
            headers = ['平台', '出货渠道', '不在义乌出的原因', '后续是否可以推进义乌出',
                       '推进不了的原因', '是否报关', '求和项:计划采购量', '计划采购量占比',
                       '求和项:采购总额', '采购金额占比']
            for c, h in enumerate(headers, 1): ws.cell(row=1, column=c).value = h
            row_num = 2
            for item in hierarchy:
                lvl = item['level']
                if lvl == 'platform': ws.cell(row=row_num, column=1).value = 'Amazon'
                elif lvl == 'channel': ws.cell(row=row_num, column=2).value = item['channel']
                elif lvl == 'reason_m': ws.cell(row=row_num, column=3).value = item['reason_m']
                elif lvl == 'reason_n': ws.cell(row=row_num, column=4).value = item['reason_n']
                elif lvl == 'reason_o': ws.cell(row=row_num, column=5).value = item['reason_o']
                elif lvl == 'customs': ws.cell(row=row_num, column=6).value = item['customs']
                elif lvl == 'total': ws.cell(row=row_num, column=1).value = '总计'
                ws.cell(row=row_num, column=7).value = item['qty']
                ws.cell(row=row_num, column=9).value = item['amt']
                row_num += 1
            total_row = row_num - 1
            for r in range(2, total_row + 1):
                ws.cell(row=r, column=8).value = f'=G{r}/G${total_row}'
                ws.cell(row=r, column=10).value = f'=I{r}/I${total_row}'
            reason_rows = [(idx, item['qty'], item['reason_m'])
                           for idx, item in enumerate(hierarchy)
                           if item['level'] == 'reason_m' and item['reason_m'] not in ('(空白)', '')]
            reason_rows.sort(key=lambda x: x[1], reverse=True)
            top3 = set(x[2] for x in reason_rows[:3])
            top3_idx = set(idx for idx, item in enumerate(hierarchy)
                           if item['level'] == 'reason_m' and item['reason_m'] in top3)
            yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            light_green = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
            cornflower_blue = PatternFill(start_color='30C0B4', end_color='30C0B4', fill_type='solid')
            orange_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                  top=Side(style='thin'), bottom=Side(style='thin'))
            bold_font = Font(bold=True, size=11, name='宋体')
            normal_font = Font(size=11, name='宋体')
            center_align = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 50
            for c in range(1, 11):
                cell = ws.cell(row=1, column=c)
                cell.fill = yellow_fill; cell.font = bold_font; cell.alignment = center_align
                if enable_border: cell.border = thin_border
            for r in range(2, total_row + 1):
                lvl = hierarchy[r - 2]['level']
                for c in range(1, 11):
                    cell = ws.cell(row=r, column=c)
                    cell.font = normal_font
                    if enable_border: cell.border = thin_border
                    cell.alignment = center_align
                ws.cell(row=r, column=8).number_format = '0.00%'
                ws.cell(row=r, column=10).number_format = '0.00%'
                if lvl == 'platform': pass
                elif lvl in ('channel', 'total'):
                    for c in range(1, 11): ws.cell(row=r, column=c).fill = light_green
                elif lvl == 'reason_m' and (r - 2) in top3_idx:
                    for c in range(1, 11): ws.cell(row=r, column=c).fill = orange_fill
                elif lvl == 'customs' and hierarchy[r - 2]['customs'] == '是':
                    for c in range(1, 11): ws.cell(row=r, column=c).fill = cornflower_blue
            for c, w in enumerate([12, 12, 18, 22, 18, 12, 18, 15, 18, 15], 1):
                ws.column_dimensions[get_column_letter(c)].width = w
            log(f"任务5: {sn} 创建完成 ({total_row}行)")
        
        # 保存
        tmp_path = output_file + '.tmp'
        wb.save(tmp_path); wb.close()
        time.sleep(0.5)
        shutil.copy2(tmp_path, output_file); os.remove(tmp_path)
        
        elapsed = time.time() - start_time
        log(f"\n全部任务完成! 耗时: {elapsed:.1f}秒")
        log(f"输出文件: {output_file}")
        return True
    
    except Exception as e:
        import traceback
        log(f"\n❌ 处理异常: {e}")
        log(traceback.format_exc())
        return False
