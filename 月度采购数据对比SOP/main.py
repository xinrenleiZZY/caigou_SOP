"""
主程序 - 月度采购数据对比SOP 整合脚本
集成任务1-4，自动检测最新月份，从源数据计算所有数值
"""
import sys, os, re, shutil, time, zipfile, tempfile
from copy import copy
from openpyxl import load_workbook, Workbook

# ========== 导入配置 ==========
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
try:
    from config import YEAR1, YEAR2, SOURCE_FILE, OUTPUT_FILE, ENABLE_BORDER
except ImportError:
    YEAR1, YEAR2 = 25, 26
    SOURCE_FILE = r'e:\ZY2026\采购需求-月度采购数据对比SOP\（IT试验版初版）26.6.11--25.5和26.5采购数据对比.xlsx'
    OUTPUT_FILE = r'e:\ZY2026\采购需求-月度采购数据对比SOP\（IT试验版初版）26.6.11--25.5和26.5采购数据对比-整合输出-v6.xlsx'
    ENABLE_BORDER = False

def log(msg):
    print(f"[{time.strftime('%H:%M:%S')}] {msg}")

# ====================================================================
# 第一部分：检测最新月份
# ====================================================================
def detect_latest_common_month(source_file, year1, year2):
    """检测两个年份共有的最新月份（如25.5和26.5都存在时取5）"""
    wb = load_workbook(source_file, read_only=True, data_only=True)
    sheet_names = set(wb.sheetnames)
    wb.close()
    
    pattern1 = re.compile(rf'^{re.escape(str(year1))}\.(\d+)$')
    pattern2 = re.compile(rf'^{re.escape(str(year2))}\.(\d+)$')
    
    months1 = set()
    months2 = set()
    for s in sheet_names:
        m = pattern1.match(s)
        if m:
            months1.add(int(m.group(1)))
        m = pattern2.match(s)
        if m:
            months2.add(int(m.group(1)))
    
    common = months1 & months2
    if not common:
        # 回退：各自取最新
        m1 = max(months1) if months1 else None
        m2 = max(months2) if months2 else None
        log(f"警告: 未找到{year1}和{year2}共有的月份, year1最新={m1}, year2最新={m2}")
        return m1, m2
    
    max_month = max(common)
    log(f"共有月份: {sorted(common)}, 取最新: {max_month}")
    return max_month, max_month

log(f"配置: YEAR1={YEAR1}, YEAR2={YEAR2}, 源文件={SOURCE_FILE}")
log("正在检测最新月份...")
MONTH1, MONTH2 = detect_latest_common_month(SOURCE_FILE, YEAR1, YEAR2)

if MONTH1 is None or MONTH2 is None:
    log(f"错误: 未找到{YEAR1}年或{YEAR2}年的数据工作表")
    sys.exit(1)

log(f"检测到最新月份: {YEAR1}.{MONTH1}, {YEAR2}.{MONTH2}")

# 构建工作表名称
SHEET1 = f"{YEAR1}.{MONTH1}"
SHEET2 = f"{YEAR2}.{MONTH2}"
SHEET_CH_AMAZON = f"{YEAR2}年出货渠道对比-Amazon"
SHEET_CH_COMPARE = f"{YEAR1}年、{YEAR2}年出货渠道对比"
SHEET_3PLATFORM = f"{YEAR1}年、{YEAR2}年采购量和采购总额对比-三平台"
SHEET_CUSTOMS = f"{YEAR1}年、{YEAR2}年报关占比表--Amazon"

SHEETS_KEEP = [SHEET1, SHEET2, SHEET_CH_AMAZON, SHEET_CH_COMPARE, SHEET_3PLATFORM, SHEET_CUSTOMS]

# ====================================================================
# 第二部分：复制文件并保留指定工作表，修复autofilter
# ====================================================================
def fix_autofilter(file_path):
    """移除xlsx中的autofilter元素"""
    tdir = tempfile.mkdtemp()
    with zipfile.ZipFile(file_path, 'r') as z:
        z.extractall(tdir)
    fixed = 0
    for root, dirs, files in os.walk(tdir):
        for f in files:
            fp = os.path.join(root, f)
            try:
                content = open(fp, 'rb').read().decode('utf-8')
            except:
                continue
            if '<autoFilter' in content:
                old = content
                content = re.sub(r'<autoFilter[^>]*>.*?</autoFilter>', '', content, flags=re.DOTALL)
                content = re.sub(r'<autoFilter[^>]*/>', '', content)
                if content != old:
                    open(fp, 'wb').write(content.encode('utf-8'))
                    fixed += 1
    os.remove(file_path)
    with zipfile.ZipFile(file_path, 'w', zipfile.ZIP_DEFLATED) as z:
        for root, dirs, files in os.walk(tdir):
            for f in files:
                z.write(os.path.join(root, f), os.path.relpath(os.path.join(root, f), tdir))
    shutil.rmtree(tdir)
    return fixed

def copy_and_clean(source, output, keep_sheets):
    """复制源文件、修复autofilter、只保留指定工作表"""
    shutil.copy2(source, output)
    log("已复制源文件")
    
    # 先修复autofilter，否则openpyxl无法读取
    fixed = fix_autofilter(output)
    log(f"已修复 {fixed} 个autofilter")
    
    # 再用openpyxl删掉不需要的工作表
    wb = load_workbook(output)
    sheets_to_remove = [s for s in wb.sheetnames if s not in keep_sheets]
    for s in sheets_to_remove:
        del wb[s]
    
    wb.save(output)
    wb.close()
    log(f"已清理工作表，保留 {len(keep_sheets)} 个: {keep_sheets}")
    
    return output

log("正在复制并清理工作簿...")
OUTPUT = copy_and_clean(SOURCE_FILE, OUTPUT_FILE, SHEETS_KEEP)

# ====================================================================
# 第三部分：读取源数据并通过计算得到所有需要的数据
# ====================================================================
def calculate_all_data(source_file, sheet1, sheet2):
    """从源数据工作表计算所有平台/渠道/报关数据"""
    wb = load_workbook(source_file, read_only=True, data_only=True)
    
    result = {}
    for sn in [sheet1, sheet2]:
        ws = wb[sn]
        platform_data = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            platform = str(row[3]).strip() if row[3] else ''
            if platform not in ['Amazon', 'temu', 'jit']:
                continue
            qty = row[7] or 0
            amt = row[10] if row[10] else 0
            channel = str(row[11]).strip() if row[11] else ''
            is_customs = str(row[15]).strip() if row[15] else ''
            
            if platform not in platform_data:
                platform_data[platform] = {'total_qty': 0, 'total_amt': 0,
                                           'gz_qty': 0, 'gz_amt': 0,
                                           'yw_qty': 0, 'yw_amt': 0,
                                           'customs_qty': 0, 'customs_amt': 0}
            pd = platform_data[platform]
            pd['total_qty'] += qty
            pd['total_amt'] += amt
            if channel == '广州':
                pd['gz_qty'] += qty; pd['gz_amt'] += amt
            elif channel == '义乌':
                pd['yw_qty'] += qty; pd['yw_amt'] += amt
            if is_customs == '是':
                pd['customs_qty'] += qty; pd['customs_amt'] += amt
        
        result[sn] = platform_data
    
    wb.close()
    return result

def calculate_channel_data(source_file, sheet1, sheet2):
    """计算出货渠道数据（更简单的API）"""
    wb = load_workbook(source_file, read_only=True, data_only=True)
    result = {}
    for sn in [sheet1, sheet2]:
        ws = wb[sn]
        gz_qty, gz_amt, yw_qty, yw_amt = 0, 0, 0, 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            p = str(row[3]).strip() if row[3] else ''
            if p != 'Amazon':
                continue
            qty = row[7] or 0
            amt = row[10] or 0
            ch = str(row[11]).strip() if row[11] else ''
            if ch == '广州':
                gz_qty += qty; gz_amt += amt
            elif ch == '义乌':
                yw_qty += qty; yw_amt += amt
        total_qty = gz_qty + yw_qty
        total_amt = gz_amt + yw_amt
        result[sn] = {
            'gz_qty': gz_qty, 'gz_amt': round(gz_amt, 2),
            'yw_qty': yw_qty, 'yw_amt': round(yw_amt, 2),
            'total_qty': total_qty, 'total_amt': round(total_amt, 2),
        }
    wb.close()
    return result

log("正在计算源数据...")
all_data = calculate_all_data(SOURCE_FILE, SHEET1, SHEET2)
channel_data = calculate_channel_data(SOURCE_FILE, SHEET1, SHEET2)

# 提取Amazon各渠道数据
ch1 = channel_data[SHEET1]  # 25.5的渠道数据
ch2 = channel_data[SHEET2]  # 26.5的渠道数据
log(f"{SHEET1} Amazon: 广州={ch1['gz_qty']}, 义乌={ch1['yw_qty']}, 合计={ch1['total_qty']}")
log(f"{SHEET2} Amazon: 广州={ch2['gz_qty']}, 义乌={ch2['yw_qty']}, 合计={ch2['total_qty']}")

# 提取三平台数据
amz_pd1 = all_data[SHEET1].get('Amazon', {})
amz_pd2 = all_data[SHEET2].get('Amazon', {})
temu_pd1 = all_data[SHEET1].get('temu', {})
temu_pd2 = all_data[SHEET2].get('temu', {})
jit_pd1 = all_data[SHEET1].get('jit', {})
jit_pd2 = all_data[SHEET2].get('jit', {})

# 提取Amazon报关数据
customs1 = {'total_qty': amz_pd1.get('total_qty', 0), 'total_amt': amz_pd1.get('total_amt', 0),
            'customs_qty': amz_pd1.get('customs_qty', 0), 'customs_amt': amz_pd1.get('customs_amt', 0)}
customs2 = {'total_qty': amz_pd2.get('total_qty', 0), 'total_amt': amz_pd2.get('total_amt', 0),
            'customs_qty': amz_pd2.get('customs_qty', 0), 'customs_amt': amz_pd2.get('customs_amt', 0)}
log(f"Amazon报关: {SHEET1}({customs1['customs_qty']}/{customs1['total_qty']}), {SHEET2}({customs2['customs_qty']}/{customs2['total_qty']})")

# ====================================================================
# 第四部分：任务1 - 26年出货渠道对比-Amazon 追加月度数据
# ====================================================================
def copy_style(src, dst):
    if src.has_style:
        dst.font = copy(src.font)
        dst.border = copy(src.border)
        dst.fill = copy(src.fill)
        dst.number_format = copy(src.number_format)
        dst.alignment = copy(src.alignment)

def task1_channel_amazon(wb, ch_data, month):
    """追加出货渠道数据"""
    ws = wb[SHEET_CH_AMAZON]
    last_row = ws.max_row
    gz_row = last_row + 1
    yw_row = last_row + 2
    
    # 从已有数据行复制格式
    for c in range(1, 9):
        copy_style(ws.cell(row=8, column=c), ws.cell(row=gz_row, column=c))
        copy_style(ws.cell(row=9, column=c), ws.cell(row=yw_row, column=c))
    
    # 广州行
    ws.cell(row=gz_row, column=1).value = month
    ws.cell(row=gz_row, column=2).value = '广州'
    ws.cell(row=gz_row, column=3).value = ch_data['gz_qty']
    ws.cell(row=gz_row, column=4).value = f'=C{gz_row}/$E${gz_row}'
    ws.cell(row=gz_row, column=5).value = f'=C{gz_row}+C{yw_row}'
    ws.cell(row=gz_row, column=6).value = ch_data['gz_amt']
    ws.cell(row=gz_row, column=7).value = f'=F{gz_row}/$H${gz_row}'
    ws.cell(row=gz_row, column=8).value = f'=F{gz_row}+F{yw_row}'
    
    # 义乌行
    ws.cell(row=yw_row, column=2).value = '义乌'
    ws.cell(row=yw_row, column=3).value = ch_data['yw_qty']
    ws.cell(row=yw_row, column=4).value = f'=C{yw_row}/$E${gz_row}'
    ws.cell(row=yw_row, column=6).value = ch_data['yw_amt']
    ws.cell(row=yw_row, column=7).value = f'=F{yw_row}/$H${gz_row}'
    
    # 合并单元格 (A, E, H)
    ws.merge_cells(start_row=gz_row, start_column=1, end_row=yw_row, end_column=1)
    ws.merge_cells(start_row=gz_row, start_column=5, end_row=yw_row, end_column=5)
    ws.merge_cells(start_row=gz_row, start_column=8, end_row=yw_row, end_column=8)
    
    from openpyxl.styles import Alignment
    align = Alignment(horizontal='center', vertical='center')
    ws.cell(row=gz_row, column=1).alignment = align
    ws.cell(row=gz_row, column=5).alignment = align
    ws.cell(row=gz_row, column=8).alignment = align
    
    log(f"任务1: 已追加{SHEET_CH_AMAZON} 月份{month} 数据")

# ====================================================================
# 第五部分：任务2 - 25年、26年出货渠道对比 追加最新月份列
# ====================================================================
def get_next_channel_columns(ws):
    """动态计算下一组列号：每月份占2列(值+占比)，从B-C开始"""
    from openpyxl.utils import get_column_letter
    col = 2  # B列开始
    while True:
        val = ws.cell(row=1, column=col).value
        if val is not None and str(val).replace('.', '').isdigit():
            col += 2  # 跳过这组2列
        else:
            break
    return col, col + 1  # (值列, 占比列)

def task2_channel_compare(wb, ch1_data, ch2_data, month1, month2):
    """动态追加出货渠道对比列"""
    ws = wb[SHEET_CH_COMPARE]
    
    val_col, pct_col = get_next_channel_columns(ws)
    from openpyxl.utils import get_column_letter
    col_letter_val = get_column_letter(val_col)
    col_letter_pct = get_column_letter(pct_col)
    log(f"  动态列: {col_letter_val}{val_col}(值) / {col_letter_pct}{pct_col}(占比)")
    
    ws.column_dimensions[col_letter_val].width = ws.column_dimensions['B'].width or 9
    ws.column_dimensions[col_letter_pct].width = ws.column_dimensions['C'].width or 9
    
    # 复制格式模板
    for src_r, dst_r in [(1,1),(2,2),(3,3),(4,4),(7,7),(8,8),(9,9),(10,10),(13,13),(14,14),(15,15)]:
        copy_style(ws.cell(row=src_r, column=2), ws.cell(row=dst_r, column=val_col))
        copy_style(ws.cell(row=src_r, column=3), ws.cell(row=dst_r, column=pct_col))
    
    from openpyxl.styles import Alignment
    align = Alignment(horizontal='center', vertical='center')
    
    # --- 25年 section (rows 1-4) ---
    ws.merge_cells(start_row=1, start_column=val_col, end_row=1, end_column=pct_col)
    ws.cell(row=1, column=val_col).value = f'{YEAR1}.{month1}'
    ws.cell(row=1, column=val_col).alignment = align
    ws.cell(row=2, column=val_col).value = ch1_data['gz_qty']
    ws.cell(row=2, column=pct_col).value = f'={col_letter_val}2/{col_letter_val}4'
    ws.cell(row=3, column=val_col).value = ch1_data['yw_qty']
    ws.cell(row=3, column=pct_col).value = f'={col_letter_val}3/{col_letter_val}4'
    ws.cell(row=4, column=val_col).value = f'={col_letter_val}2+{col_letter_val}3'
    ws.cell(row=4, column=pct_col).value = f'={col_letter_val}4/{col_letter_val}4'
    
    # --- 26年 section (rows 7-10) ---
    ws.merge_cells(start_row=7, start_column=val_col, end_row=7, end_column=pct_col)
    ws.cell(row=7, column=val_col).value = f'{YEAR2}.{month2}'
    ws.cell(row=7, column=val_col).alignment = align
    ws.cell(row=8, column=val_col).value = ch2_data['gz_qty']
    ws.cell(row=8, column=pct_col).value = f'={col_letter_val}8/{col_letter_val}$10'
    ws.cell(row=9, column=val_col).value = ch2_data['yw_qty']
    ws.cell(row=9, column=pct_col).value = f'={col_letter_val}9/{col_letter_val}$10'
    ws.cell(row=10, column=val_col).value = f'={col_letter_val}8+{col_letter_val}9'
    ws.cell(row=10, column=pct_col).value = f'={col_letter_val}10/{col_letter_val}$10'
    
    # --- 差异值 section (rows 13-15) ---
    ws.merge_cells(start_row=13, start_column=val_col, end_row=13, end_column=pct_col)
    ws.cell(row=13, column=val_col).value = month2
    ws.cell(row=13, column=val_col).alignment = align
    ws.cell(row=14, column=val_col).value = f'={col_letter_val}8-{col_letter_val}2'
    ws.cell(row=14, column=pct_col).value = f'={col_letter_pct}8-{col_letter_pct}2'
    ws.cell(row=15, column=val_col).value = f'={col_letter_val}9-{col_letter_val}3'
    ws.cell(row=15, column=pct_col).value = f'={col_letter_pct}9-{col_letter_pct}3'
    
    log(f"任务2: 已追加{SHEET_CH_COMPARE} {col_letter_val}-{col_letter_pct}列 ({YEAR1}.{month1}/{YEAR2}.{month2})")

# ====================================================================
# 第六部分：任务3 - 三平台工作表追加最新月份列
# ====================================================================
def get_next_platform_column(ws):
    """动态计算下一列号：每月份占1列，从B列开始"""
    from openpyxl.utils import get_column_letter
    col = 2  # B列
    while True:
        val = ws.cell(row=3, column=col).value  # 25年header行
        if val is not None:
            col += 1
        else:
            break
    return col

def task3_three_platforms(wb, all_data, month1, month2):
    """动态追加三平台数据列"""
    ws = wb[SHEET_3PLATFORM]
    
    new_col = get_next_platform_column(ws)
    from openpyxl.utils import get_column_letter
    col_letter = get_column_letter(new_col)
    log(f"  动态列: {col_letter}{new_col}")
    
    ws.column_dimensions[col_letter].width = ws.column_dimensions['B'].width or 12
    
    platform_positions = {'Amazon': 1, 'temu': 18, 'jit': 35}
    
    for platform, start in platform_positions.items():
        pd1 = all_data.get(SHEET1, {}).get(platform, {})
        pd2 = all_data.get(SHEET2, {}).get(platform, {})
        
        q1 = pd1.get('total_qty', 0); a1 = round(pd1.get('total_amt', 0), 2)
        q2 = pd2.get('total_qty', 0); a2 = round(pd2.get('total_amt', 0), 2)
        qd = q2 - q1; ad = round(a2 - a1, 2)
        
        for src_r, dst_r in [(start+2, start+2),(start+3, start+3),(start+4, start+4),
                              (start+7, start+7),(start+8, start+8),(start+9, start+9),
                              (start+12, start+12),(start+13, start+13),(start+14, start+14)]:
            copy_style(ws.cell(row=src_r, column=2), ws.cell(row=dst_r, column=new_col))
        
        ws.cell(row=start+2, column=new_col).value = f'{YEAR1}.{month1}'
        ws.cell(row=start+3, column=new_col).value = q1
        ws.cell(row=start+4, column=new_col).value = a1
        ws.cell(row=start+7, column=new_col).value = f'{YEAR2}.{month2}'
        ws.cell(row=start+8, column=new_col).value = q2
        ws.cell(row=start+9, column=new_col).value = a2
        ws.cell(row=start+12, column=new_col).value = month2
        ws.cell(row=start+13, column=new_col).value = qd
        ws.cell(row=start+14, column=new_col).value = ad
        
        log(f"任务3: {platform} - {YEAR1}.{month1}(qty={q1}, amt={a1}), "
            f"{YEAR2}.{month2}(qty={q2}, amt={a2}), diff(qty={qd}, amt={ad})")

# ====================================================================
# 第七部分：任务4 - 报关占比表插入月度数据
# ====================================================================
def adjust_all_formulas(ws, insert_at, count=2):
    """insert_rows后调整公式行号引用"""
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if v and isinstance(v, str) and v.startswith('='):
                def repl(m):
                    pre = m.group(1)
                    rn = m.group(2)
                    dollar = '$' if rn.startswith('$') else ''
                    clean = rn.lstrip('$')
                    num = int(clean)
                    if num >= insert_at:
                        return f'{pre}{dollar}{num + count}'
                    return m.group(0)
                cell.value = re.sub(r'(\$?[A-Z]+\$?)(\d+)', repl, v)

def task4_customs_table(wb, customs1, customs2, month1, month2):
    """插入报关数据行"""
    ws = wb[SHEET_CUSTOMS]
    
    # 查找2025{month1-1:02d}和2026{month2-1:02d}的结束位置
    def find_month_end(ws, month_val):
        """找到月份数据块的结束位置（是行+1，即插入位置）"""
        for r in range(1, ws.max_row + 1):
            if ws.cell(row=r, column=1).value == month_val:
                # month_val出现在两行（total行+报关行），返回报关行后的位置
                return r + 2
        return None
    
    # 插入25.5 (2025{month1:02d})
    m1_label = int(f'{2000+YEAR1}{month1:02d}')
    m2_label = int(f'{2000+YEAR2}{month2:02d}')
    
    # 找到插入位置
    pos1 = find_month_end(ws, int(f'{2000+YEAR1}{month1-1:02d}'))
    pos2 = find_month_end(ws, int(f'{2000+YEAR2}{month2-1:02d}'))
    
    if pos1 is None:
        log(f"错误: 未找到{YEAR1}年{month1-1}月的位置")
        pos1 = 14  # fallback
    
    if pos2 is None:
        log(f"错误: 未找到{YEAR2}年{month2-1}月的位置")
        # pos2 = pos1 + 16  # approximate
    
    def insert_row_pair(ws, ins, label, c_data):
        ws.insert_rows(ins, 2)
        adjust_all_formulas(ws, ins, 2)
        
        # 逐列复制格式：C列(采购量->General), D列(采购总额->#,##0.00), E列(占比->0.00%), F列(占比->0.00%)
        template_rows = [13, 14]  # 202504 total行 和 报关行
        
        # Total行
        copy_style(ws.cell(row=template_rows[0], column=1), ws.cell(row=ins, column=1))
        copy_style(ws.cell(row=template_rows[0], column=2), ws.cell(row=ins, column=2))
        copy_style(ws.cell(row=template_rows[0], column=3), ws.cell(row=ins, column=3))
        copy_style(ws.cell(row=template_rows[0], column=4), ws.cell(row=ins, column=4))
        copy_style(ws.cell(row=template_rows[0], column=5), ws.cell(row=ins, column=5))
        copy_style(ws.cell(row=template_rows[0], column=6), ws.cell(row=ins, column=6))
        ws.cell(row=ins, column=1).value = label
        ws.cell(row=ins, column=3).value = c_data['total_qty']
        ws.cell(row=ins, column=4).value = round(c_data['total_amt'], 2)
        
        # 报关行
        copy_style(ws.cell(row=template_rows[1], column=1), ws.cell(row=ins+1, column=1))
        copy_style(ws.cell(row=template_rows[1], column=2), ws.cell(row=ins+1, column=2))
        copy_style(ws.cell(row=template_rows[1], column=3), ws.cell(row=ins+1, column=3))
        copy_style(ws.cell(row=template_rows[1], column=4), ws.cell(row=ins+1, column=4))
        copy_style(ws.cell(row=template_rows[1], column=5), ws.cell(row=ins+1, column=5))
        copy_style(ws.cell(row=template_rows[1], column=6), ws.cell(row=ins+1, column=6))
        ws.cell(row=ins+1, column=1).value = label
        ws.cell(row=ins+1, column=2).value = '是'
        ws.cell(row=ins+1, column=3).value = c_data['customs_qty']
        ws.cell(row=ins+1, column=4).value = round(c_data['customs_amt'], 2)
        ws.cell(row=ins+1, column=5).value = f'=C{ins+1}/C{ins}'
        ws.cell(row=ins+1, column=6).value = f'=D{ins+1}/D{ins}'
        
        log(f"任务4: 插入 {label} 行 {ins}-{ins+1}")
    
    insert_row_pair(ws, pos1, m1_label, customs1)
    insert_row_pair(ws, pos2 + 2, m2_label, customs2)  # +2因为有第一次插入

# ====================================================================
# 第七部分-B：任务5 - 数据透视分析表
# ====================================================================
from openpyxl.styles import PatternFill, Font as OpenFont, Alignment as OpenAlign, Border as OpenBorder, Side as OpenSide

def build_amazon_hierarchy(ws, max_row):
    """从源表读取Amazon数据，按层级分组汇总
    层级: Amazon总行 → 出货渠道 → 不在义乌出的原因 → 后续推进 → 推进原因 → 报关
    """
    # 读取所有Amazon行
    rows = []
    for r in range(2, max_row + 1):
        platform = ws.cell(row=r, column=4).value
        if platform and str(platform).strip() == 'Amazon':
            qty = ws.cell(row=r, column=8).value or 0
            amt = ws.cell(row=r, column=11).value or 0
            channel = str(ws.cell(row=r, column=12).value or '').strip()
            reason_m = str(ws.cell(row=r, column=13).value or '').strip()
            reason_n = str(ws.cell(row=r, column=14).value or '').strip()
            reason_o = str(ws.cell(row=r, column=15).value or '').strip()
            customs = str(ws.cell(row=r, column=16).value or '').strip()
            rows.append({'qty': qty, 'amt': amt, 'channel': channel,
                         'reason_m': reason_m, 'reason_n': reason_n,
                         'reason_o': reason_o, 'customs': customs})

    total_qty = sum(r['qty'] for r in rows)
    total_amt = sum(r['amt'] for r in rows)

    result = []

    # === Amazon总行（第一行） ===
    result.append({'level': 'platform', 'channel': '', 'reason_m': '',
                   'reason_n': '', 'reason_o': '', 'customs': '',
                   'qty': total_qty, 'amt': round(total_amt, 2)})

    channels = {}
    for row in rows:
        ch = row['channel']
        channels.setdefault(ch, []).append(row)

    for ch_name in ['广州', '义乌']:
        if ch_name not in channels:
            continue
        ch_rows = channels[ch_name]
        ch_qty = sum(r['qty'] for r in ch_rows)
        ch_amt = sum(r['amt'] for r in ch_rows)

        result.append({'level': 'channel', 'channel': ch_name,
                       'reason_m': '', 'reason_n': '', 'reason_o': '', 'customs': '',
                       'qty': ch_qty, 'amt': ch_amt})

        reasons_m = {}
        for row in ch_rows:
            rm = row['reason_m'] if row['reason_m'] else '(空白)'
            reasons_m.setdefault(rm, []).append(row)

        for rm_name, rm_rows in sorted(reasons_m.items()):
            rm_qty = sum(r['qty'] for r in rm_rows)
            rm_amt = sum(r['amt'] for r in rm_rows)
            result.append({'level': 'reason_m', 'channel': '', 'reason_m': rm_name,
                           'reason_n': '', 'reason_o': '', 'customs': '',
                           'qty': rm_qty, 'amt': rm_amt})

            reasons_n = {}
            for row in rm_rows:
                rn = row['reason_n'] if row['reason_n'] else '(空白)'
                reasons_n.setdefault(rn, []).append(row)

            for rn_name, rn_rows in sorted(reasons_n.items()):
                rn_qty = sum(r['qty'] for r in rn_rows)
                rn_amt = sum(r['amt'] for r in rn_rows)
                result.append({'level': 'reason_n', 'channel': '', 'reason_m': '',
                               'reason_n': rn_name, 'reason_o': '', 'customs': '',
                               'qty': rn_qty, 'amt': rn_amt})

                reasons_o = {}
                for row in rn_rows:
                    ro = row['reason_o'] if row['reason_o'] else '(空白)'
                    reasons_o.setdefault(ro, []).append(row)

                for ro_name, ro_rows in sorted(reasons_o.items()):
                    ro_qty = sum(r['qty'] for r in ro_rows)
                    ro_amt = sum(r['amt'] for r in ro_rows)
                    result.append({'level': 'reason_o', 'channel': '', 'reason_m': '',
                                   'reason_n': '', 'reason_o': ro_name, 'customs': '',
                                   'qty': ro_qty, 'amt': ro_amt})

                    customs_map = {}
                    for row in ro_rows:
                        c = row['customs'] if row['customs'] else '(空白)'
                        customs_map.setdefault(c, []).append(row)

                    for c_name, c_rows in sorted(customs_map.items()):
                        c_qty = sum(r['qty'] for r in c_rows)
                        c_amt = sum(r['amt'] for r in c_rows)
                        result.append({'level': 'customs', 'channel': '', 'reason_m': '',
                                       'reason_n': '', 'reason_o': '', 'customs': c_name,
                                       'qty': c_qty, 'amt': c_amt})

    result.append({'level': 'total', 'channel': '总计', 'reason_m': '',
                   'reason_n': '', 'reason_o': '', 'customs': '',
                   'qty': total_qty, 'amt': round(total_amt, 2)})
    return result

def task5_create_analysis(wb, sheet_name, source_sheet_name):
    """创建分析工作表"""
    ws_src = wb[source_sheet_name]
    hierarchy = build_amazon_hierarchy(ws_src, ws_src.max_row)

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)

    headers = ['平台', '出货渠道', '不在义乌出的原因', '后续是否可以推进义乌出',
               '推进不了的原因', '是否报关', '求和项:计划采购量', '计划采购量占比',
               '求和项:采购总额', '采购金额占比']
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c).value = h

    row_num = 2
    for item in hierarchy:
        lvl = item['level']
        if lvl == 'platform':
            ws.cell(row=row_num, column=1).value = 'Amazon'
        elif lvl == 'channel':
            ws.cell(row=row_num, column=2).value = item['channel']
        elif lvl == 'reason_m':
            ws.cell(row=row_num, column=3).value = item['reason_m']
        elif lvl == 'reason_n':
            ws.cell(row=row_num, column=4).value = item['reason_n']
        elif lvl == 'reason_o':
            ws.cell(row=row_num, column=5).value = item['reason_o']
        elif lvl == 'customs':
            ws.cell(row=row_num, column=6).value = item['customs']
        elif lvl == 'total':
            ws.cell(row=row_num, column=1).value = '总计'
        ws.cell(row=row_num, column=7).value = item['qty']
        ws.cell(row=row_num, column=9).value = item['amt']
        row_num += 1

    total_row = row_num - 1

    for r in range(2, total_row + 1):
        ws.cell(row=r, column=8).value = f'=G{r}/G${total_row}'
        ws.cell(row=r, column=10).value = f'=I{r}/I${total_row}'

    # 计算所有reason_m行（排除空白原因），按采购量取TOP3
    reason_m_rows_by_qty = []
    for idx, item in enumerate(hierarchy):
        if item['level'] == 'reason_m' and item['reason_m'] not in ('(空白)', ''):
            reason_m_rows_by_qty.append((idx, item['qty'], item['reason_m']))
    reason_m_rows_by_qty.sort(key=lambda x: x[1], reverse=True)
    top3_reasons = set(x[2] for x in reason_m_rows_by_qty[:3])
    top3_reason_indices = set()
    if top3_reasons:
        for idx, item in enumerate(hierarchy):
            if item['level'] == 'reason_m' and item['reason_m'] in top3_reasons:
                top3_reason_indices.add(idx)

    # 格式化
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    light_green = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
    cornflower_blue = PatternFill(start_color='30C0B4', end_color='30C0B4', fill_type='solid')
    orange_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    thin_border = OpenBorder(left=OpenSide(style='thin'), right=OpenSide(style='thin'),
                              top=OpenSide(style='thin'), bottom=OpenSide(style='thin'))
    bold_font = OpenFont(bold=True, size=11, name='宋体')
    normal_font = OpenFont(size=11, name='宋体')
    center_align = OpenAlign(horizontal='center', vertical='center')

    # 表头 - 黄色 + 行高50
    ws.row_dimensions[1].height = 50
    for c in range(1, 11):
        cell = ws.cell(row=1, column=c)
        cell.fill = yellow_fill
        cell.font = bold_font
        cell.alignment = center_align
        if ENABLE_BORDER:
            cell.border = thin_border

    for r in range(2, total_row + 1):
        lvl = hierarchy[r - 2]['level']
        for c in range(1, 11):
            cell = ws.cell(row=r, column=c)
            cell.font = normal_font
            if ENABLE_BORDER:
                cell.border = thin_border
            cell.alignment = center_align
        ws.cell(row=r, column=8).number_format = '0.00%'
        ws.cell(row=r, column=10).number_format = '0.00%'

        if lvl == 'platform':
            pass  # Amazon总行不上色
        elif lvl in ('channel', 'total'):
            for c in range(1, 11):
                ws.cell(row=r, column=c).fill = light_green
        elif lvl == 'reason_m' and (r - 2) in top3_reason_indices:
            for c in range(1, 11):
                ws.cell(row=r, column=c).fill = orange_fill
        elif lvl == 'customs' and hierarchy[r - 2]['customs'] == '是':
            for c in range(1, 11):
                ws.cell(row=r, column=c).fill = cornflower_blue

    from openpyxl.utils import get_column_letter
    for c, w in enumerate([12, 12, 18, 22, 18, 12, 18, 15, 18, 15], 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    log(f"任务5: 创建 {sheet_name} ({total_row}行, TOP3原因={top3_reasons})")

# ====================================================================
# 第八部分：执行所有任务
# ====================================================================
log("\n开始执行所有任务...")
wb = load_workbook(OUTPUT)

log("\n--- 任务1: 出货渠道对比 ---")
task1_channel_amazon(wb, ch2, MONTH2)

log("\n--- 任务2: 出货渠道对比J-K列 ---")
task2_channel_compare(wb, ch1, ch2, MONTH1, MONTH2)

log("\n--- 任务3: 三平台数据 ---")
task3_three_platforms(wb, all_data, MONTH1, MONTH2)

log("\n--- 任务4: 报关占比表 ---")
task4_customs_table(wb, customs1, customs2, MONTH1, MONTH2)

log("\n--- 任务5: 数据分析透视表 ---")
task5_create_analysis(wb, f'{YEAR1}.{MONTH1}数据分析', SHEET1)
task5_create_analysis(wb, f'{YEAR2}.{MONTH2}数据分析', SHEET2)

# ====================================================================
# 第九部分：保存并验证
# ====================================================================
tmp_path = OUTPUT + '.final_tmp'
wb.save(tmp_path)
wb.close()

time.sleep(1)
shutil.copy2(tmp_path, OUTPUT)
os.remove(tmp_path)

log(f"\n全部任务完成！输出文件: {OUTPUT}")

# 快速验证
log("正在验证...")
wb2 = load_workbook(OUTPUT)
log(f"最终工作表: {wb2.sheetnames}")
for s in wb2.sheetnames:
    ws = wb2[s]
    log(f"  {s}: {ws.max_row}行 x {ws.max_column}列")
wb2.close()
log("验证完成！")
