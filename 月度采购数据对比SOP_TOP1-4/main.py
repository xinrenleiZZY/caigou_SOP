"""
主程序 - 月度采购数据对比SOP 整合脚本
集成任务1-4，自动检测最新月份，从源数据计算所有数值
"""
import sys, os, re, shutil, time, zipfile, tempfile
from copy import copy
from openpyxl import load_workbook, Workbook

# ========== 导入配置 ==========
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
try:
    from config import YEAR1, YEAR2, SOURCE_FILE, OUTPUT_FILE
except ImportError:
    YEAR1, YEAR2 = 25, 26
    SOURCE_FILE = r'e:\ZY2026\采购需求-月度采购数据对比SOP\（IT试验版初版）26.6.11--25.5和26.5采购数据对比.xlsx'
    OUTPUT_FILE = r'e:\ZY2026\采购需求-月度采购数据对比SOP\（IT试验版初版）26.6.11--25.5和26.5采购数据对比-整合输出-v2.xlsx'

def log(msg):
    print(f"[{time.strftime('%H:%M:%S')}] {msg}")

# ====================================================================
# 第一部分：检测最新月份
# ====================================================================
def detect_latest_common_month(source_file, year1, year2):
    """检测两个年份共有的最新月份（如25.5和26.5都存在时取5）"""
    wb = load_workbook(source_file, read_only=True, data_only=True)
    sheet_names = set(wb.sheetnames)
    wb.close()
    
    pattern1 = re.compile(rf'^{re.escape(str(year1))}\.(\d+)$')
    pattern2 = re.compile(rf'^{re.escape(str(year2))}\.(\d+)$')
    
    months1 = set()
    months2 = set()
    for s in sheet_names:
        m = pattern1.match(s)
        if m:
            months1.add(int(m.group(1)))
        m = pattern2.match(s)
        if m:
            months2.add(int(m.group(1)))
    
    common = months1 & months2
    if not common:
        # 回退：各自取最新
        m1 = max(months1) if months1 else None
        m2 = max(months2) if months2 else None
        log(f"警告: 未找到{year1}和{year2}共有的月份, year1最新={m1}, year2最新={m2}")
        return m1, m2
    
    max_month = max(common)
    log(f"共有月份: {sorted(common)}, 取最新: {max_month}")
    return max_month, max_month

log(f"配置: YEAR1={YEAR1}, YEAR2={YEAR2}, 源文件={SOURCE_FILE}")
log("正在检测最新月份...")
MONTH1, MONTH2 = detect_latest_common_month(SOURCE_FILE, YEAR1, YEAR2)

if MONTH1 is None or MONTH2 is None:
    log(f"错误: 未找到{YEAR1}年或{YEAR2}年的数据工作表")
    sys.exit(1)

log(f"检测到最新月份: {YEAR1}.{MONTH1}, {YEAR2}.{MONTH2}")

# 构建工作表名称
SHEET1 = f"{YEAR1}.{MONTH1}"
SHEET2 = f"{YEAR2}.{MONTH2}"
SHEET_CH_AMAZON = f"{YEAR2}年出货渠道对比-Amazon"
SHEET_CH_COMPARE = f"{YEAR1}年、{YEAR2}年出货渠道对比"
SHEET_3PLATFORM = f"{YEAR1}年、{YEAR2}年采购量和采购总额对比-三平台"
SHEET_CUSTOMS = f"{YEAR1}年、{YEAR2}年报关占比表--Amazon"

SHEETS_KEEP = [SHEET1, SHEET2, SHEET_CH_AMAZON, SHEET_CH_COMPARE, SHEET_3PLATFORM, SHEET_CUSTOMS]

# ====================================================================
# 第二部分：复制文件并保留指定工作表，修复autofilter
# ====================================================================
def fix_autofilter(file_path):
    """移除xlsx中的autofilter元素"""
    tdir = tempfile.mkdtemp()
    with zipfile.ZipFile(file_path, 'r') as z:
        z.extractall(tdir)
    fixed = 0
    for root, dirs, files in os.walk(tdir):
        for f in files:
            fp = os.path.join(root, f)
            try:
                content = open(fp, 'rb').read().decode('utf-8')
            except:
                continue
            if '<autoFilter' in content:
                old = content
                content = re.sub(r'<autoFilter[^>]*>.*?</autoFilter>', '', content, flags=re.DOTALL)
                content = re.sub(r'<autoFilter[^>]*/>', '', content)
                if content != old:
                    open(fp, 'wb').write(content.encode('utf-8'))
                    fixed += 1
    os.remove(file_path)
    with zipfile.ZipFile(file_path, 'w', zipfile.ZIP_DEFLATED) as z:
        for root, dirs, files in os.walk(tdir):
            for f in files:
                z.write(os.path.join(root, f), os.path.relpath(os.path.join(root, f), tdir))
    shutil.rmtree(tdir)
    return fixed

def copy_and_clean(source, output, keep_sheets):
    """复制源文件、修复autofilter、只保留指定工作表"""
    shutil.copy2(source, output)
    log("已复制源文件")
    
    # 先修复autofilter，否则openpyxl无法读取
    fixed = fix_autofilter(output)
    log(f"已修复 {fixed} 个autofilter")
    
    # 再用openpyxl删掉不需要的工作表
    wb = load_workbook(output)
    sheets_to_remove = [s for s in wb.sheetnames if s not in keep_sheets]
    for s in sheets_to_remove:
        del wb[s]
    
    wb.save(output)
    wb.close()
    log(f"已清理工作表，保留 {len(keep_sheets)} 个: {keep_sheets}")
    
    return output

log("正在复制并清理工作簿...")
OUTPUT = copy_and_clean(SOURCE_FILE, OUTPUT_FILE, SHEETS_KEEP)

# ====================================================================
# 第三部分：读取源数据并通过计算得到所有需要的数据
# ====================================================================
def calculate_all_data(source_file, sheet1, sheet2):
    """从源数据工作表计算所有平台/渠道/报关数据"""
    wb = load_workbook(source_file, read_only=True, data_only=True)
    
    result = {}
    for sn in [sheet1, sheet2]:
        ws = wb[sn]
        platform_data = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            platform = str(row[3]).strip() if row[3] else ''
            if platform not in ['Amazon', 'temu', 'jit']:
                continue
            qty = row[7] or 0
            amt = row[10] if row[10] else 0
            channel = str(row[11]).strip() if row[11] else ''
            is_customs = str(row[15]).strip() if row[15] else ''
            
            if platform not in platform_data:
                platform_data[platform] = {'total_qty': 0, 'total_amt': 0,
                                           'gz_qty': 0, 'gz_amt': 0,
                                           'yw_qty': 0, 'yw_amt': 0,
                                           'customs_qty': 0, 'customs_amt': 0}
            pd = platform_data[platform]
            pd['total_qty'] += qty
            pd['total_amt'] += amt
            if channel == '广州':
                pd['gz_qty'] += qty; pd['gz_amt'] += amt
            elif channel == '义乌':
                pd['yw_qty'] += qty; pd['yw_amt'] += amt
            if is_customs == '是':
                pd['customs_qty'] += qty; pd['customs_amt'] += amt
        
        result[sn] = platform_data
    
    wb.close()
    return result

def calculate_channel_data(source_file, sheet1, sheet2):
    """计算出货渠道数据（更简单的API）"""
    wb = load_workbook(source_file, read_only=True, data_only=True)
    result = {}
    for sn in [sheet1, sheet2]:
        ws = wb[sn]
        gz_qty, gz_amt, yw_qty, yw_amt = 0, 0, 0, 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            p = str(row[3]).strip() if row[3] else ''
            if p != 'Amazon':
                continue
            qty = row[7] or 0
            amt = row[10] or 0
            ch = str(row[11]).strip() if row[11] else ''
            if ch == '广州':
                gz_qty += qty; gz_amt += amt
            elif ch == '义乌':
                yw_qty += qty; yw_amt += amt
        total_qty = gz_qty + yw_qty
        total_amt = gz_amt + yw_amt
        result[sn] = {
            'gz_qty': gz_qty, 'gz_amt': round(gz_amt, 2),
            'yw_qty': yw_qty, 'yw_amt': round(yw_amt, 2),
            'total_qty': total_qty, 'total_amt': round(total_amt, 2),
        }
    wb.close()
    return result

log("正在计算源数据...")
all_data = calculate_all_data(SOURCE_FILE, SHEET1, SHEET2)
channel_data = calculate_channel_data(SOURCE_FILE, SHEET1, SHEET2)

# 提取Amazon各渠道数据
ch1 = channel_data[SHEET1]  # 25.5的渠道数据
ch2 = channel_data[SHEET2]  # 26.5的渠道数据
log(f"{SHEET1} Amazon: 广州={ch1['gz_qty']}, 义乌={ch1['yw_qty']}, 合计={ch1['total_qty']}")
log(f"{SHEET2} Amazon: 广州={ch2['gz_qty']}, 义乌={ch2['yw_qty']}, 合计={ch2['total_qty']}")

# 提取三平台数据
amz_pd1 = all_data[SHEET1].get('Amazon', {})
amz_pd2 = all_data[SHEET2].get('Amazon', {})
temu_pd1 = all_data[SHEET1].get('temu', {})
temu_pd2 = all_data[SHEET2].get('temu', {})
jit_pd1 = all_data[SHEET1].get('jit', {})
jit_pd2 = all_data[SHEET2].get('jit', {})

# 提取Amazon报关数据
customs1 = {'total_qty': amz_pd1.get('total_qty', 0), 'total_amt': amz_pd1.get('total_amt', 0),
            'customs_qty': amz_pd1.get('customs_qty', 0), 'customs_amt': amz_pd1.get('customs_amt', 0)}
customs2 = {'total_qty': amz_pd2.get('total_qty', 0), 'total_amt': amz_pd2.get('total_amt', 0),
            'customs_qty': amz_pd2.get('customs_qty', 0), 'customs_amt': amz_pd2.get('customs_amt', 0)}
log(f"Amazon报关: {SHEET1}({customs1['customs_qty']}/{customs1['total_qty']}), {SHEET2}({customs2['customs_qty']}/{customs2['total_qty']})")

# ====================================================================
# 第四部分：任务1 - 26年出货渠道对比-Amazon 追加月度数据
# ====================================================================
def copy_style(src, dst):
    if src.has_style:
        dst.font = copy(src.font)
        dst.border = copy(src.border)
        dst.fill = copy(src.fill)
        dst.number_format = copy(src.number_format)
        dst.alignment = copy(src.alignment)

def task1_channel_amazon(wb, ch_data, month):
    """追加出货渠道数据"""
    ws = wb[SHEET_CH_AMAZON]
    last_row = ws.max_row
    gz_row = last_row + 1
    yw_row = last_row + 2
    
    # 从已有数据行复制格式
    for c in range(1, 9):
        copy_style(ws.cell(row=8, column=c), ws.cell(row=gz_row, column=c))
        copy_style(ws.cell(row=9, column=c), ws.cell(row=yw_row, column=c))
    
    # 广州行
    ws.cell(row=gz_row, column=1).value = month
    ws.cell(row=gz_row, column=2).value = '广州'
    ws.cell(row=gz_row, column=3).value = ch_data['gz_qty']
    ws.cell(row=gz_row, column=4).value = f'=C{gz_row}/$E${gz_row}'
    ws.cell(row=gz_row, column=5).value = f'=C{gz_row}+C{yw_row}'
    ws.cell(row=gz_row, column=6).value = ch_data['gz_amt']
    ws.cell(row=gz_row, column=7).value = f'=F{gz_row}/$H${gz_row}'
    ws.cell(row=gz_row, column=8).value = f'=F{gz_row}+F{yw_row}'
    
    # 义乌行
    ws.cell(row=yw_row, column=2).value = '义乌'
    ws.cell(row=yw_row, column=3).value = ch_data['yw_qty']
    ws.cell(row=yw_row, column=4).value = f'=C{yw_row}/$E${gz_row}'
    ws.cell(row=yw_row, column=6).value = ch_data['yw_amt']
    ws.cell(row=yw_row, column=7).value = f'=F{yw_row}/$H${gz_row}'
    
    # 合并单元格 (A, E, H)
    ws.merge_cells(start_row=gz_row, start_column=1, end_row=yw_row, end_column=1)
    ws.merge_cells(start_row=gz_row, start_column=5, end_row=yw_row, end_column=5)
    ws.merge_cells(start_row=gz_row, start_column=8, end_row=yw_row, end_column=8)
    
    from openpyxl.styles import Alignment
    align = Alignment(horizontal='center', vertical='center')
    ws.cell(row=gz_row, column=1).alignment = align
    ws.cell(row=gz_row, column=5).alignment = align
    ws.cell(row=gz_row, column=8).alignment = align
    
    log(f"任务1: 已追加{SHEET_CH_AMAZON} 月份{month} 数据")

# ====================================================================
# 第五部分：任务2 - 25年、26年出货渠道对比 追加J-K列
# ====================================================================
def task2_channel_compare(wb, ch1_data, ch2_data, month1, month2):
    """追加出货渠道对比J-K列"""
    ws = wb[SHEET_CH_COMPARE]
    
    # 设置列宽
    ws.column_dimensions['J'].width = ws.column_dimensions['B'].width or 9
    ws.column_dimensions['K'].width = ws.column_dimensions['C'].width or 9
    
    # 复制格式模板 - J列从B列(值), K列从C列(占比=0.00%)
    for src_r, dst_r in [(1,1),(2,2),(3,3),(4,4),(7,7),(8,8),(9,9),(10,10),(13,13),(14,14),(15,15)]:
        copy_style(ws.cell(row=src_r, column=2), ws.cell(row=dst_r, column=10))  # J列 = B列格式
        copy_style(ws.cell(row=src_r, column=3), ws.cell(row=dst_r, column=11))  # K列 = C列格式(0.00%)
    
    from openpyxl.styles import Alignment
    align = Alignment(horizontal='center', vertical='center')
    
    # --- 25年 section ---
    ws.merge_cells('J1:K1')
    ws.cell(row=1, column=10).value = f'{YEAR1}.{month1}'
    ws.cell(row=1, column=10).alignment = align
    ws.cell(row=2, column=10).value = ch1_data['gz_qty']
    ws.cell(row=2, column=11).value = f'=J2/J4'
    ws.cell(row=3, column=10).value = ch1_data['yw_qty']
    ws.cell(row=3, column=11).value = f'=J3/J4'
    ws.cell(row=4, column=10).value = f'=J2+J3'
    ws.cell(row=4, column=11).value = f'=J4/J4'
    
    # --- 26年 section ---
    ws.merge_cells('J7:K7')
    ws.cell(row=7, column=10).value = f'{YEAR2}.{month2}'
    ws.cell(row=7, column=10).alignment = align
    ws.cell(row=8, column=10).value = ch2_data['gz_qty']
    ws.cell(row=8, column=11).value = f'=J8/J$10'
    ws.cell(row=9, column=10).value = ch2_data['yw_qty']
    ws.cell(row=9, column=11).value = f'=J9/J$10'
    ws.cell(row=10, column=10).value = f'=J8+J9'
    ws.cell(row=10, column=11).value = f'=J10/J$10'
    
    # --- 差异值 section ---
    ws.merge_cells('J13:K13')
    ws.cell(row=13, column=10).value = month2
    ws.cell(row=13, column=10).alignment = align
    ws.cell(row=14, column=10).value = f'=J8-J2'
    ws.cell(row=14, column=11).value = f'=K8-K2'
    ws.cell(row=15, column=10).value = f'=J9-J3'
    ws.cell(row=15, column=11).value = f'=K9-K3'
    
    log(f"任务2: 已追加{SHEET_CH_COMPARE} J-K列 ({YEAR1}.{month1}/{YEAR2}.{month2})")

# ====================================================================
# 第六部分：任务3 - 三平台工作表追加G列
# ====================================================================
def task3_three_platforms(wb, all_data, month1, month2):
    """追加三平台G列数据"""
    ws = wb[SHEET_3PLATFORM]
    ws.column_dimensions['G'].width = ws.column_dimensions['B'].width or 12
    
    platform_positions = {'Amazon': 1, 'temu': 18, 'jit': 35}
    
    for platform, start in platform_positions.items():
        pd1 = all_data.get(SHEET1, {}).get(platform, {})
        pd2 = all_data.get(SHEET2, {}).get(platform, {})
        
        q1 = pd1.get('total_qty', 0); a1 = round(pd1.get('total_amt', 0), 2)
        q2 = pd2.get('total_qty', 0); a2 = round(pd2.get('total_amt', 0), 2)
        qd = q2 - q1; ad = round(a2 - a1, 2)
        
        for src_r, dst_r in [(start+2, start+2),(start+3, start+3),(start+4, start+4),
                              (start+7, start+7),(start+8, start+8),(start+9, start+9),
                              (start+12, start+12),(start+13, start+13),(start+14, start+14)]:
            copy_style(ws.cell(row=src_r, column=2), ws.cell(row=dst_r, column=7))
        
        col = 7
        ws.cell(row=start+2, column=col).value = f'{YEAR1}.{month1}'
        ws.cell(row=start+3, column=col).value = q1
        ws.cell(row=start+4, column=col).value = a1
        ws.cell(row=start+7, column=col).value = f'{YEAR2}.{month2}'
        ws.cell(row=start+8, column=col).value = q2
        ws.cell(row=start+9, column=col).value = a2
        ws.cell(row=start+12, column=col).value = month2
        ws.cell(row=start+13, column=col).value = qd
        ws.cell(row=start+14, column=col).value = ad
        
        log(f"任务3: {platform} - {YEAR1}.{month1}(qty={q1}, amt={a1}), "
            f"{YEAR2}.{month2}(qty={q2}, amt={a2}), diff(qty={qd}, amt={ad})")

# ====================================================================
# 第七部分：任务4 - 报关占比表插入月度数据
# ====================================================================
def adjust_all_formulas(ws, insert_at, count=2):
    """insert_rows后调整公式行号引用"""
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if v and isinstance(v, str) and v.startswith('='):
                def repl(m):
                    pre = m.group(1)
                    rn = m.group(2)
                    dollar = '$' if rn.startswith('$') else ''
                    clean = rn.lstrip('$')
                    num = int(clean)
                    if num >= insert_at:
                        return f'{pre}{dollar}{num + count}'
                    return m.group(0)
                cell.value = re.sub(r'(\$?[A-Z]+\$?)(\d+)', repl, v)

def task4_customs_table(wb, customs1, customs2, month1, month2):
    """插入报关数据行"""
    ws = wb[SHEET_CUSTOMS]
    
    # 查找2025{month1-1:02d}和2026{month2-1:02d}的结束位置
    def find_month_end(ws, month_val):
        """找到月份数据块的结束位置（是行+1，即插入位置）"""
        for r in range(1, ws.max_row + 1):
            if ws.cell(row=r, column=1).value == month_val:
                # month_val出现在两行（total行+报关行），返回报关行后的位置
                return r + 2
        return None
    
    # 插入25.5 (2025{month1:02d})
    m1_label = int(f'{2000+YEAR1}{month1:02d}')
    m2_label = int(f'{2000+YEAR2}{month2:02d}')
    
    # 找到插入位置
    pos1 = find_month_end(ws, int(f'{2000+YEAR1}{month1-1:02d}'))
    pos2 = find_month_end(ws, int(f'{2000+YEAR2}{month2-1:02d}'))
    
    if pos1 is None:
        log(f"错误: 未找到{YEAR1}年{month1-1}月的位置")
        pos1 = 14  # fallback
    
    if pos2 is None:
        log(f"错误: 未找到{YEAR2}年{month2-1}月的位置")
        # pos2 = pos1 + 16  # approximate
    
    def insert_row_pair(ws, ins, label, c_data):
        ws.insert_rows(ins, 2)
        adjust_all_formulas(ws, ins, 2)
        
        # 逐列复制格式：C列(采购量->General), D列(采购总额->#,##0.00), E列(占比->0.00%), F列(占比->0.00%)
        template_rows = [13, 14]  # 202504 total行 和 报关行
        
        # Total行
        copy_style(ws.cell(row=template_rows[0], column=1), ws.cell(row=ins, column=1))
        copy_style(ws.cell(row=template_rows[0], column=2), ws.cell(row=ins, column=2))
        copy_style(ws.cell(row=template_rows[0], column=3), ws.cell(row=ins, column=3))
        copy_style(ws.cell(row=template_rows[0], column=4), ws.cell(row=ins, column=4))
        copy_style(ws.cell(row=template_rows[0], column=5), ws.cell(row=ins, column=5))
        copy_style(ws.cell(row=template_rows[0], column=6), ws.cell(row=ins, column=6))
        ws.cell(row=ins, column=1).value = label
        ws.cell(row=ins, column=3).value = c_data['total_qty']
        ws.cell(row=ins, column=4).value = round(c_data['total_amt'], 2)
        
        # 报关行
        copy_style(ws.cell(row=template_rows[1], column=1), ws.cell(row=ins+1, column=1))
        copy_style(ws.cell(row=template_rows[1], column=2), ws.cell(row=ins+1, column=2))
        copy_style(ws.cell(row=template_rows[1], column=3), ws.cell(row=ins+1, column=3))
        copy_style(ws.cell(row=template_rows[1], column=4), ws.cell(row=ins+1, column=4))
        copy_style(ws.cell(row=template_rows[1], column=5), ws.cell(row=ins+1, column=5))
        copy_style(ws.cell(row=template_rows[1], column=6), ws.cell(row=ins+1, column=6))
        ws.cell(row=ins+1, column=1).value = label
        ws.cell(row=ins+1, column=2).value = '是'
        ws.cell(row=ins+1, column=3).value = c_data['customs_qty']
        ws.cell(row=ins+1, column=4).value = round(c_data['customs_amt'], 2)
        ws.cell(row=ins+1, column=5).value = f'=C{ins+1}/C{ins}'
        ws.cell(row=ins+1, column=6).value = f'=D{ins+1}/D{ins}'
        
        log(f"任务4: 插入 {label} 行 {ins}-{ins+1}")
    
    insert_row_pair(ws, pos1, m1_label, customs1)
    insert_row_pair(ws, pos2 + 2, m2_label, customs2)  # +2因为有第一次插入

# ====================================================================
# 第八部分：执行所有任务
# ====================================================================
log("\n开始执行所有任务...")
wb = load_workbook(OUTPUT)

log("\n--- 任务1: 出货渠道对比 ---")
task1_channel_amazon(wb, ch2, MONTH2)

log("\n--- 任务2: 出货渠道对比J-K列 ---")
task2_channel_compare(wb, ch1, ch2, MONTH1, MONTH2)

log("\n--- 任务3: 三平台数据 ---")
task3_three_platforms(wb, all_data, MONTH1, MONTH2)

log("\n--- 任务4: 报关占比表 ---")
task4_customs_table(wb, customs1, customs2, MONTH1, MONTH2)

# ====================================================================
# 第九部分：保存并验证
# ====================================================================
tmp_path = OUTPUT + '.final_tmp'
wb.save(tmp_path)
wb.close()

time.sleep(1)
shutil.copy2(tmp_path, OUTPUT)
os.remove(tmp_path)

log(f"\n全部任务完成！输出文件: {OUTPUT}")

# 快速验证
log("正在验证...")
wb2 = load_workbook(OUTPUT)
log(f"最终工作表: {wb2.sheetnames}")
for s in wb2.sheetnames:
    ws = wb2[s]
    log(f"  {s}: {ws.max_row}行 x {ws.max_column}列")
wb2.close()
log("验证完成！")
